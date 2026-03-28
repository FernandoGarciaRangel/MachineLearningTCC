"""
Machine Learning em Dutos de Óleo e Gás
========================================
Modelos implementados (classificação e regressão):
  1. Árvore de Decisão
  2. Regressão Logística (Softmax)
  3. Random Forest
  4. Gradient Boosting (XGBoost)

Datasets:
  - dataset_com_manutencao.xlsx
  - dataset_sem_manutencao.xlsx

Autores: Fernando Garcia Rangel / Leonardo dos Santos Rodrigues
"""

import warnings
warnings.filterwarnings('ignore')

from pathlib import Path

import pandas as pd
import numpy as np
from sklearn.tree import DecisionTreeClassifier, DecisionTreeRegressor
from sklearn.linear_model import LogisticRegression
from sklearn.ensemble import (RandomForestClassifier, RandomForestRegressor,
                               GradientBoostingClassifier, GradientBoostingRegressor)
from sklearn.metrics import (confusion_matrix, classification_report,
                              accuracy_score, f1_score,
                              mean_absolute_error, mean_squared_error, r2_score)
from sklearn.preprocessing import StandardScaler
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── NOTA: Para usar XGBoost real, descomente as linhas abaixo e
#          comente as linhas de GradientBoosting do sklearn acima.
# from xgboost import XGBClassifier, XGBRegressor

# =============================================================================
# CONFIGURAÇÕES
# =============================================================================

FEATURES = [
    'ano',
    'posicao_m',
    'dist_sold_ant_m',
    'compr_tubo_m',
    'ie_enc',
    'tipo_enc',
    'pos_horaria_graus',
    'esp_mm',
    'compr_mm',
    'larg_mm',
    'prof_pct',
    'erf',
    'tipo_pof_enc'
]

TARGET_CLASS = 'classe_risco_enc'
TARGET_REG   = 'taxa_corrosao_ano'

CLASSES_NOMES = ['Baixo', 'Medio', 'Alto']

_PASTA_TRATADOS = Path(__file__).resolve().parent.parent / 'Tratados'
DATASETS = {
    'Com Manutencao': _PASTA_TRATADOS / 'dataset_com_manutencao.xlsx',
    'Sem Manutencao': _PASTA_TRATADOS / 'dataset_sem_manutencao.xlsx',
}

RANDOM_STATE = 42

# =============================================================================
# CARREGAMENTO DOS DADOS
# =============================================================================

def carregar_splits(filepath):
    """Carrega treino, validação e teste de um arquivo Excel."""
    treino    = pd.read_excel(filepath, sheet_name='treino')
    validacao = pd.read_excel(filepath, sheet_name='validacao')
    teste     = pd.read_excel(filepath, sheet_name='teste')

    # Treino + validação combinados para treino final
    treino_val = pd.concat([treino, validacao], ignore_index=True)

    return treino, validacao, teste, treino_val


def preparar_xy(df, target):
    X = df[FEATURES].copy()
    y = df[target].copy()
    return X, y

# =============================================================================
# DEFINIÇÃO DOS MODELOS
# =============================================================================

def get_modelos_classificacao():
    return {
        'Arvore de Decisao': DecisionTreeClassifier(
            max_depth=6,
            min_samples_split=10,
            random_state=RANDOM_STATE
        ),
        'Regressao Logistica': LogisticRegression(
            solver='lbfgs',
            max_iter=1000,
            random_state=RANDOM_STATE
        ),
        'Random Forest': RandomForestClassifier(
            n_estimators=200,
            max_depth=8,
            min_samples_split=5,
            random_state=RANDOM_STATE,
            n_jobs=-1
        ),
        # Gradient Boosting (sklearn) — substitua por XGBClassifier se disponível:
        # 'Gradient Boosting': XGBClassifier(
        #     n_estimators=200, learning_rate=0.1, max_depth=5,
        #     use_label_encoder=False, eval_metric='mlogloss',
        #     random_state=RANDOM_STATE
        # )
        'Gradient Boosting': GradientBoostingClassifier(
            n_estimators=200,
            learning_rate=0.1,
            max_depth=5,
            random_state=RANDOM_STATE
        )
    }


def get_modelos_regressao():
    return {
        'Arvore de Decisao': DecisionTreeRegressor(
            max_depth=6,
            min_samples_split=10,
            random_state=RANDOM_STATE
        ),
        'Regressao Logistica': LogisticRegression(
            solver='lbfgs',
            max_iter=1000,
            random_state=RANDOM_STATE
        ),  # placeholder — veja nota abaixo
        'Random Forest': RandomForestRegressor(
            n_estimators=200,
            max_depth=8,
            min_samples_split=5,
            random_state=RANDOM_STATE,
            n_jobs=-1
        ),
        # Gradient Boosting (sklearn) — substitua por XGBRegressor se disponível:
        # 'Gradient Boosting': XGBRegressor(
        #     n_estimators=200, learning_rate=0.1, max_depth=5,
        #     random_state=RANDOM_STATE
        # )
        'Gradient Boosting': GradientBoostingRegressor(
            n_estimators=200,
            learning_rate=0.1,
            max_depth=5,
            random_state=RANDOM_STATE
        )
    }

# =============================================================================
# TREINAMENTO E AVALIAÇÃO
# =============================================================================

def treinar_classificacao(nome_modelo, modelo, X_treino, y_treino,
                           X_val, y_val, X_teste, y_teste, classes_presentes):
    """Treina e avalia um modelo de classificação."""
    # Escalonamento para Regressão Logística
    if 'Logistica' in nome_modelo:
        scaler = StandardScaler()
        X_treino = scaler.fit_transform(X_treino)
        X_val    = scaler.transform(X_val)
        X_teste  = scaler.transform(X_teste)

    modelo.fit(X_treino, y_treino)

    y_pred_val   = modelo.predict(X_val)
    y_pred_teste = modelo.predict(X_teste)

    # Métricas no conjunto de validação
    acc_val = accuracy_score(y_val, y_pred_val)
    f1_val  = f1_score(y_val, y_pred_val, average='weighted', zero_division=0)

    # Métricas no conjunto de teste
    acc_teste = accuracy_score(y_teste, y_pred_teste)
    f1_teste  = f1_score(y_teste, y_pred_teste, average='weighted', zero_division=0)

    # Matriz de confusão (teste)
    cm = confusion_matrix(y_teste, y_pred_teste, labels=list(range(len(CLASSES_NOMES))))

    # Relatório detalhado (teste)
    nomes_presentes = [CLASSES_NOMES[i] for i in sorted(classes_presentes)]
    relatorio = classification_report(
        y_teste, y_pred_teste,
        labels=sorted(classes_presentes),
        target_names=nomes_presentes,
        zero_division=0,
        output_dict=True
    )

    return {
        'acc_val':    round(acc_val, 4),
        'f1_val':     round(f1_val, 4),
        'acc_teste':  round(acc_teste, 4),
        'f1_teste':   round(f1_teste, 4),
        'cm':         cm,
        'relatorio':  relatorio,
        'y_pred':     y_pred_teste,
        'y_real':     y_teste.values
    }


def treinar_regressao(nome_modelo, modelo, X_treino, y_treino,
                      X_val, y_val, X_teste, y_teste):
    """Treina e avalia um modelo de regressão."""
    # Regressão Logística não faz regressão contínua —
    # para este target usamos Árvore como substituto e registramos N/A
    if 'Logistica' in nome_modelo:
        return None

    if 'Logistica' in nome_modelo:
        scaler = StandardScaler()
        X_treino = scaler.fit_transform(X_treino)
        X_val    = scaler.transform(X_val)
        X_teste  = scaler.transform(X_teste)

    modelo.fit(X_treino, y_treino)

    y_pred_val   = modelo.predict(X_val)
    y_pred_teste = modelo.predict(X_teste)

    mae_val  = mean_absolute_error(y_val, y_pred_val)
    rmse_val = np.sqrt(mean_squared_error(y_val, y_pred_val))
    r2_val   = r2_score(y_val, y_pred_val)

    mae_teste  = mean_absolute_error(y_teste, y_pred_teste)
    rmse_teste = np.sqrt(mean_squared_error(y_teste, y_pred_teste))
    r2_teste   = r2_score(y_teste, y_pred_teste)

    return {
        'mae_val':   round(mae_val, 4),
        'rmse_val':  round(rmse_val, 4),
        'r2_val':    round(r2_val, 4),
        'mae_teste': round(mae_teste, 4),
        'rmse_teste':round(rmse_teste, 4),
        'r2_teste':  round(r2_teste, 4),
        'y_pred':    y_pred_teste,
        'y_real':    y_teste.values
    }

# =============================================================================
# EXECUÇÃO PRINCIPAL
# =============================================================================

todos_resultados = {}  # armazena tudo para o Excel

for nome_ds, arquivo in DATASETS.items():
    print(f"\n{'#'*60}")
    print(f"  DATASET: {nome_ds}")
    print(f"{'#'*60}")

    treino, validacao, teste, treino_val = carregar_splits(arquivo)

    # Classes presentes no dataset
    classes_presentes = set(treino_val[TARGET_CLASS].unique()) | \
                        set(teste[TARGET_CLASS].unique())

    X_treino_val, y_treino_val_c = preparar_xy(treino_val, TARGET_CLASS)
    X_val,  y_val_c  = preparar_xy(validacao, TARGET_CLASS)
    X_teste, y_teste_c = preparar_xy(teste, TARGET_CLASS)

    tem_regressao = TARGET_REG in treino_val.columns
    if tem_regressao:
        X_treino_val_r, y_treino_val_r = preparar_xy(treino_val, TARGET_REG)
        X_val_r, y_val_r = preparar_xy(validacao, TARGET_REG)
        X_teste_r, y_teste_r = preparar_xy(teste, TARGET_REG)

    resultados_ds = {'classificacao': {}, 'regressao': {}}

    # ── CLASSIFICAÇÃO ──────────────────────────────────────────────────────
    print(f"\n  [ CLASSIFICAÇÃO ]")
    for nome_m, modelo in get_modelos_classificacao().items():
        res = treinar_classificacao(
            nome_m, modelo,
            X_treino_val, y_treino_val_c,
            X_val, y_val_c,
            X_teste, y_teste_c,
            classes_presentes
        )
        resultados_ds['classificacao'][nome_m] = res
        print(f"    {nome_m:25s} | Acurácia: {res['acc_teste']:.4f} | F1: {res['f1_teste']:.4f}")

    # ── REGRESSÃO ──────────────────────────────────────────────────────────
    if tem_regressao:
        print(f"\n  [ REGRESSÃO ]")
        for nome_m, modelo in get_modelos_regressao().items():
            res = treinar_regressao(
                nome_m, modelo,
                X_treino_val_r, y_treino_val_r,
                X_val_r, y_val_r,
                X_teste_r, y_teste_r
            )
            if res is None:
                print(f"    {'Regressao Logistica':25s} | N/A (não aplicável para regressão contínua)")
                resultados_ds['regressao']['Regressao Logistica'] = None
            else:
                resultados_ds['regressao'][nome_m] = res
                print(f"    {nome_m:25s} | MAE: {res['mae_teste']:.4f} | RMSE: {res['rmse_teste']:.4f} | R²: {res['r2_teste']:.4f}")
    else:
        print(f"\n  [ REGRESSÃO ] — omitida (coluna '{TARGET_REG}' ausente no Excel)")

    todos_resultados[nome_ds] = resultados_ds

# =============================================================================
# EXPORTAR RESULTADOS PARA EXCEL
# =============================================================================

print("\n\nGerando arquivo de resultados...")

wb = Workbook()
wb.remove(wb.active)  # remove aba padrão

# ── estilos ───────────────────────────────────────────────────────────────────
COR_HEADER    = "1F4E79"   # azul escuro
COR_SUBHEADER = "2E75B6"   # azul médio
COR_LENTA     = "C6EFCE"   # verde claro
COR_MODERADA  = "FFEB9C"   # amarelo claro
COR_ACELERADA = "FFC7CE"   # vermelho claro
COR_DATASET1  = "D6E4F0"
COR_DATASET2  = "FDEBD0"

def estilo_header(cell, cor=COR_HEADER):
    cell.font      = Font(bold=True, color="FFFFFF", size=11)
    cell.fill      = PatternFill("solid", fgColor=cor)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def borda_fina(cell):
    lado = Side(style='thin', color='BFBFBF')
    cell.border = Border(left=lado, right=lado, top=lado, bottom=lado)

def formatar_bloco(ws, linha_ini, col_ini, linhas, colunas):
    for r in range(linha_ini, linha_ini + linhas):
        for c in range(col_ini, col_ini + colunas):
            borda_fina(ws.cell(r, c))


# ══════════════════════════════════════════════════════════════════════════════
# ABA 1: RESUMO GERAL
# ══════════════════════════════════════════════════════════════════════════════
ws_res = wb.create_sheet("Resumo Geral")

ws_res.column_dimensions['A'].width = 22
ws_res.column_dimensions['B'].width = 26
for col in ['C','D','E','F','G','H']:
    ws_res.column_dimensions[col].width = 14

# Título
ws_res.merge_cells('A1:H1')
ws_res['A1'] = 'Machine Learning em Dutos de Óleo e Gás — Resumo de Resultados'
ws_res['A1'].font      = Font(bold=True, size=14, color="FFFFFF")
ws_res['A1'].fill      = PatternFill("solid", fgColor=COR_HEADER)
ws_res['A1'].alignment = Alignment(horizontal="center", vertical="center")
ws_res.row_dimensions[1].height = 30

linha = 3

for nome_ds, cor_ds in [('Com Manutencao', COR_DATASET1), ('Sem Manutencao', COR_DATASET2)]:

    # Subtítulo dataset
    ws_res.merge_cells(f'A{linha}:H{linha}')
    ws_res[f'A{linha}'] = f'Dataset: {nome_ds.replace("_"," ")}'
    ws_res[f'A{linha}'].font      = Font(bold=True, size=12, color="FFFFFF")
    ws_res[f'A{linha}'].fill      = PatternFill("solid", fgColor=COR_SUBHEADER)
    ws_res[f'A{linha}'].alignment = Alignment(horizontal="center")
    ws_res.row_dimensions[linha].height = 22
    linha += 1

    # ── Classificação ──────────────────────────────────────────────────────
    ws_res.merge_cells(f'A{linha}:H{linha}')
    ws_res[f'A{linha}'] = 'CLASSIFICAÇÃO'
    ws_res[f'A{linha}'].font      = Font(bold=True, size=11)
    ws_res[f'A{linha}'].fill      = PatternFill("solid", fgColor="BDD7EE")
    ws_res[f'A{linha}'].alignment = Alignment(horizontal="center")
    linha += 1

    headers_c = ['Modelo', 'Acurácia (Val)', 'F1 (Val)', 'Acurácia (Teste)', 'F1 (Teste)']
    for i, h in enumerate(headers_c):
        c = ws_res.cell(linha, i + 1, h)
        estilo_header(c, COR_SUBHEADER)
    linha += 1

    for nome_m, res in todos_resultados[nome_ds]['classificacao'].items():
        ws_res.cell(linha, 1, nome_m)
        ws_res.cell(linha, 2, res['acc_val'])
        ws_res.cell(linha, 3, res['f1_val'])
        ws_res.cell(linha, 4, res['acc_teste'])
        ws_res.cell(linha, 5, res['f1_teste'])
        for c in range(1, 6):
            borda_fina(ws_res.cell(linha, c))
            ws_res.cell(linha, c).alignment = Alignment(horizontal="center")
        linha += 1

    linha += 1

    # ── Regressão ──────────────────────────────────────────────────────────
    ws_res.merge_cells(f'A{linha}:H{linha}')
    ws_res[f'A{linha}'] = 'REGRESSÃO'
    ws_res[f'A{linha}'].font      = Font(bold=True, size=11)
    ws_res[f'A{linha}'].fill      = PatternFill("solid", fgColor="FCE4D6")
    ws_res[f'A{linha}'].alignment = Alignment(horizontal="center")
    linha += 1

    headers_r = ['Modelo', 'MAE (Val)', 'RMSE (Val)', 'R² (Val)', 'MAE (Teste)', 'RMSE (Teste)', 'R² (Teste)']
    for i, h in enumerate(headers_r):
        c = ws_res.cell(linha, i + 1, h)
        estilo_header(c, COR_SUBHEADER)
    linha += 1

    for nome_m, res in todos_resultados[nome_ds]['regressao'].items():
        ws_res.cell(linha, 1, nome_m)
        if res is None:
            ws_res.merge_cells(f'B{linha}:H{linha}')
            ws_res.cell(linha, 2, 'N/A — Regressão Logística não aplicável para target contínuo')
            ws_res.cell(linha, 2).alignment = Alignment(horizontal="center")
        else:
            ws_res.cell(linha, 2, res['mae_val'])
            ws_res.cell(linha, 3, res['rmse_val'])
            ws_res.cell(linha, 4, res['r2_val'])
            ws_res.cell(linha, 5, res['mae_teste'])
            ws_res.cell(linha, 6, res['rmse_teste'])
            ws_res.cell(linha, 7, res['r2_teste'])
        for c in range(1, 8):
            borda_fina(ws_res.cell(linha, c))
            ws_res.cell(linha, c).alignment = Alignment(horizontal="center")
        linha += 1

    linha += 2


# ══════════════════════════════════════════════════════════════════════════════
# ABAS DE MATRIZES DE CONFUSÃO
# ══════════════════════════════════════════════════════════════════════════════

for nome_ds in todos_resultados:
    ws_cm = wb.create_sheet(f"CM_{nome_ds[:3]}_Manut" if 'Com' in nome_ds else f"CM_{nome_ds[:3]}_SemMan")
    ws_cm = wb.create_sheet(f"Conf_{nome_ds.replace(' ','_')[:20]}")

    ws_cm.column_dimensions['A'].width = 26
    for col in ['B','C','D','E','F','G']:
        ws_cm.column_dimensions[col].width = 14

    # Título
    ws_cm.merge_cells('A1:F1')
    ws_cm['A1'] = f'Matrizes de Confusão — {nome_ds}'
    ws_cm['A1'].font      = Font(bold=True, size=13, color="FFFFFF")
    ws_cm['A1'].fill      = PatternFill("solid", fgColor=COR_HEADER)
    ws_cm['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_cm.row_dimensions[1].height = 28

    linha = 3
    classes_ds = sorted(todos_resultados[nome_ds]['classificacao'][
        list(todos_resultados[nome_ds]['classificacao'].keys())[0]
    ]['cm'].shape)[0]

    # detectar classes presentes
    todas_classes = CLASSES_NOMES

    for nome_m, res in todos_resultados[nome_ds]['classificacao'].items():
        cm = res['cm']

        # Subtítulo modelo
        ws_cm.merge_cells(f'A{linha}:F{linha}')
        ws_cm[f'A{linha}'] = nome_m
        ws_cm[f'A{linha}'].font      = Font(bold=True, size=11, color="FFFFFF")
        ws_cm[f'A{linha}'].fill      = PatternFill("solid", fgColor=COR_SUBHEADER)
        ws_cm[f'A{linha}'].alignment = Alignment(horizontal="center")
        ws_cm.row_dimensions[linha].height = 20
        linha += 1

        # Métricas resumidas
        ws_cm.cell(linha, 1, f"Acurácia (Teste): {res['acc_teste']:.4f}  |  F1 Ponderado: {res['f1_teste']:.4f}")
        ws_cm.cell(linha, 1).font = Font(italic=True)
        ws_cm.merge_cells(f'A{linha}:F{linha}')
        linha += 1

        # Header da matriz
        ws_cm.cell(linha, 1, 'Real \\ Previsto')
        estilo_header(ws_cm.cell(linha, 1))
        for j, cls in enumerate(todas_classes):
            c = ws_cm.cell(linha, j + 2, cls)
            estilo_header(c)
        linha += 1

        # Linhas da matriz
        cores_linha = [COR_LENTA, COR_MODERADA, COR_ACELERADA, "FFD7D7"]
        for i, cls_real in enumerate(todas_classes):
            ws_cm.cell(linha, 1, cls_real)
            ws_cm.cell(linha, 1).font = Font(bold=True)
            ws_cm.cell(linha, 1).fill = PatternFill("solid", fgColor=cores_linha[i])
            for j in range(len(todas_classes)):
                val = cm[i][j] if i < cm.shape[0] and j < cm.shape[1] else 0
                cell = ws_cm.cell(linha, j + 2, int(val))
                cell.alignment = Alignment(horizontal="center")
                # destacar diagonal (acertos)
                if i == j and val > 0:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="00B050")
                    cell.font = Font(bold=True, color="FFFFFF")
                borda_fina(cell)
            borda_fina(ws_cm.cell(linha, 1))
            linha += 1

        linha += 2


# ══════════════════════════════════════════════════════════════════════════════
# ABAS DETALHADAS POR DATASET
# ══════════════════════════════════════════════════════════════════════════════

for nome_ds in todos_resultados:
    ws_det = wb.create_sheet(f"Det_{nome_ds.replace(' ','_')[:18]}")
    ws_det.column_dimensions['A'].width = 22
    for col in ['B','C','D','E','F','G','H','I','J']:
        ws_det.column_dimensions[col].width = 15

    ws_det.merge_cells('A1:I1')
    ws_det['A1'] = f'Relatório Detalhado por Classe — {nome_ds}'
    ws_det['A1'].font      = Font(bold=True, size=13, color="FFFFFF")
    ws_det['A1'].fill      = PatternFill("solid", fgColor=COR_HEADER)
    ws_det['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_det.row_dimensions[1].height = 28

    linha = 3

    for nome_m, res in todos_resultados[nome_ds]['classificacao'].items():
        ws_det.merge_cells(f'A{linha}:I{linha}')
        ws_det[f'A{linha}'] = nome_m
        ws_det[f'A{linha}'].font      = Font(bold=True, size=11, color="FFFFFF")
        ws_det[f'A{linha}'].fill      = PatternFill("solid", fgColor=COR_SUBHEADER)
        ws_det[f'A{linha}'].alignment = Alignment(horizontal="center")
        linha += 1

        headers = ['Classe', 'Precision', 'Recall', 'F1-Score', 'Support']
        for i, h in enumerate(headers):
            c = ws_det.cell(linha, i + 1, h)
            estilo_header(c, COR_SUBHEADER)
        linha += 1

        rel = res['relatorio']
        for cls_nome in CLASSES_NOMES:
            if cls_nome in rel:
                row = rel[cls_nome]
                ws_det.cell(linha, 1, cls_nome)
                ws_det.cell(linha, 2, round(row['precision'], 4))
                ws_det.cell(linha, 3, round(row['recall'], 4))
                ws_det.cell(linha, 4, round(row['f1-score'], 4))
                ws_det.cell(linha, 5, int(row['support']))
                for c in range(1, 6):
                    borda_fina(ws_det.cell(linha, c))
                    ws_det.cell(linha, c).alignment = Alignment(horizontal="center")
                linha += 1

        # Linha de acurácia geral
        ws_det.cell(linha, 1, 'Acurácia Geral')
        ws_det.cell(linha, 1).font = Font(bold=True)
        ws_det.merge_cells(f'B{linha}:E{linha}')
        ws_det.cell(linha, 2, round(res['acc_teste'], 4))
        ws_det.cell(linha, 2).alignment = Alignment(horizontal="center")
        ws_det.cell(linha, 2).font = Font(bold=True)
        for c in range(1, 6):
            borda_fina(ws_det.cell(linha, c))
        linha += 3


# ── Salvar ────────────────────────────────────────────────────────────────────
# Remover aba vazia duplicada se existir
for nome in list(wb.sheetnames):
    if nome.startswith('CM_'):
        del wb[nome]

output_path = Path(__file__).resolve().parent / 'resultados_modelos.xlsx'
wb.save(output_path)
print(f"Resultados salvos em: {output_path}")
print("Processamento concluído!")
